import time
<<<<<<< HEAD
import pytest
=======

import pytest
import request
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
import inspect
from appium import webdriver
from appium.options.common.base import AppiumOptions
from pages import main
from selenium.common.exceptions import TimeoutException
<<<<<<< HEAD
from openpyxl import Workbook
from openpyxl.drawing.image import Image as xlImage

# Load the existing Excel workbook
wb = Workbook()

# Access the active worksheet
ws = wb.active

@pytest.fixture(scope="session")
def excel_results(request):
    # No need to adjust row and column dimensions here
=======
import allure
from openpyxl import Workbook, load_workbook
# Load the existing Excel workbook
wb = load_workbook(r"C:\Users\admin\PycharmProjects\Appium\test_results.xlsx")
ws = wb.active
# Define a fixture to save the workbook after the tests have run
@pytest.fixture(scope="session")
def excel_results(request):
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
    def fin():
        wb.save("test_results.xlsx")
        wb.close()

    request.addfinalizer(fin)
<<<<<<< HEAD

=======
# Define a fixture to open the appium driver
# @pytest.fixture(scope="module")
# def appium_driver():
#     options = AppiumOptions()
#     options.load_capabilities({
#     "platformName": "Android",
#     "deviceName": "RZ8R21SMYNN",
#     })
#     driver = webdriver.Remote("http://127.0.0.1:4723/wd/hub", options=options)
#     yield driver
#     driver.quit()
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9

@pytest.fixture(scope="module")
def appium_driver():
    options = AppiumOptions()
    options.load_capabilities({
        "platformName": "Android",
        "deviceName": "RZ8R21SMYNN",
        # Add other desired capabilities as needed
    })
<<<<<<< HEAD
=======
    # Add other desired capabilities as needed
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9

    driver = webdriver.Remote("http://127.0.0.1:4723/wd/hub", options=options)
    yield driver
    # driver.quit()


<<<<<<< HEAD
@pytest.mark.usefixtures("appium_driver", "excel_results")
class Testorderpage:
    @staticmethod
    def test_openapp_tc_01(appium_driver, excel_results):
        try:
            main.TestAppiumAutomation.test_click_docile_delivery_element(appium_driver)
            time.sleep(7)

            screenshot_name = "test_openapp_tc_01.png"
            appium_driver.save_screenshot(screenshot_name)
            # screenshot = xlImage(screenshot_name)
=======
@pytest.mark.usefixtures("appium_driver","excel_results")
class Testorderpage:
    @staticmethod
    def test_openapp_tc_01(appium_driver,excel_results):
        try:
            main.TestAppiumAutomation.test_click_docile_delivery_element(appium_driver)
            time.sleep(7)
            screenshot_name = f"test_openapp_tc_01.png"
            # allure.attach(appium_driver.get_screenshot_as_png(), name="Screenshot",
            #               attachment_type=allure.attachment_type.PNG)
            appium_driver.save_screenshot(screenshot_name)

>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
<<<<<<< HEAD
            # Append test data to the worksheet
            ws.append([f"TC-{test_id}", function_name, result,screenshot_hyperlink, "None", "None", "open the docile app",
                       "As Expected", "APK", "None"])

            # Add the screenshot image to the Excel file
            # Adjust cell position as needed

            wb.save("test_results.xlsx")

        except TimeoutException:
            result = "Fail"
            screenshot_name = "test_openapp_tc_01.png"
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]

            # Append test data to the worksheet
            ws.append([f"TC-{test_id}", function_name, result,screenshot_name, "None", "None", "open docile App",
                       "Not opening", "APK", "it's not clicking the app icon"])

          

            wb.save("test_results.xlsx")

###############################################################################################################################

=======
            ws.append(
                [f"TC-{test_id}", function_name, result, screenshot_hyperlink, "None", "None", "open the docile app",
                 "As Expected", "APK", "None"])
            wb.save("test_results.xlsx")
            wb.save("test_results.xlsx")
        except TimeoutException:
            result = "Fail"
            screenshot_name = f"test_openapp_tc_01.png"
            # allure.attach(appium_driver.get_screenshot_as_png(), name="Screenshot",
            #               attachment_type=allure.attachment_type.PNG)
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "open docile App", "Not opening", "APK", "it's not clicking the app icon"])
            wb.save("test_results.xlsx")
###############################################################################################################################
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
    def test_yestologin_tc_02(self,appium_driver):
        try:
            main.TestAppiumAutomation.test_pressyes(self,appium_driver)
            time.sleep(7)
            screenshot_name = f"test_openapp_tc_02.png"
            # allure.attach(appium_driver.get_screenshot_as_png(), name="Screenshot",
            #               attachment_type=allure.attachment_type.PNG)
            appium_driver.save_screenshot(screenshot_name)

            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append(
                [f"TC-{test_id}", function_name, result, screenshot_hyperlink, "None", "None", "YES to login",
                 "As Expected", "APK", "None"])
            wb.save("test_results.xlsx")

        except TimeoutException:
            result = "Fail"
            screenshot_name = f"test_openapp_tc_02.png"
            # allure.attach(appium_driver.get_screenshot_as_png(), name="Screenshot",
            #               attachment_type=allure.attachment_type.PNG)
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
            wb.save("test_results.xlsx")

            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "YES to login", "Not opening", "APK", "it's not clicking the app icon"])
################################################################################################################################################################
    def test_mobile_otp_tc_03(self,appium_driver):
        try:
            main.TestAppiumAutomation.test_mobile_number(self,appium_driver)
            time.sleep(7)
            screenshot_name = f"test_mobileotp_tc_03.png"
            # allure.attach(appium_driver.get_screenshot_as_png(), name="Screenshot",
            #               attachment_type=allure.attachment_type.PNG)
            appium_driver.save_screenshot(screenshot_name)

            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append(
                [f"TC-{test_id}", function_name, result, screenshot_hyperlink, "None", "None", "Login with mobile no",
                 "As Expected", "APK", "None"])
            wb.save("test_results.xlsx")


        except TimeoutException:

            result = "Fail"

            screenshot_name = f"test_mobileotp_tc_03.png"

            # allure.attach(appium_driver.get_screenshot_as_png(), name="Screenshot",
            #
            #               attachment_type=allure.attachment_type.PNG)

            appium_driver.save_screenshot(screenshot_name)

            function_name = inspect.stack()[0][3]

            test_id = function_name.split("_")[-1]

            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'

            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])

            wb.save("test_results.xlsx")

            function_name = inspect.stack()[0][3]

            test_id = function_name.split("_")[-1]

            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "Login with mobile no", "Not opening", "APK", "it's not clicking the app icon"])
############################################################################ 3 ###########################################
    def test_notification_button_tc_04(self,appium_driver):
        try:
            main.TestAppiumAutomation.test_notification_button(self,appium_driver)
            time.sleep(7)
            screenshot_name = f"test_notification_tc_04.png"
            # allure.attach(appium_driver.get_screenshot_as_png(), name="Screenshot",
            #               attachment_type=allure.attachment_type.PNG)
            appium_driver.save_screenshot(screenshot_name)

            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append(
                [f"TC-{test_id}", function_name, result, screenshot_hyperlink, "None", "None", "Click on notification",
                 "As Expected", "APK", "None"])
            wb.save("test_results.xlsx")
        except TimeoutException:
            result = "Fail"
            screenshot_name = f"test_notfication_tc_04.png"
            # allure.attach(appium_driver.get_screenshot_as_png(), name="Screenshot",
            #               attachment_type=allure.attachment_type.PNG)
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
            wb.save("test_results.xlsx")
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "Click on notification ", "Not opening", "APK", "it's not clicking the app icon"])
################################################################################################################################
<<<<<<< HEAD
    # def testorder_details_ongoing_tc_05(self,appium_driver):
    #     try:
    #         main.TestAppiumAutomation. test_dropdown_myorders_ongoing(self,appium_driver)
    #         time.sleep(7)
    #         screenshot_name = f"test_notification_tc_04.png"
    #         appium_driver.save_screenshot(screenshot_name)
    #
    #         assert True  # Example assertion
    #         result = "Pass"
    #         function_name = inspect.stack()[0][3]
    #         test_id = function_name.split("_")[-1]
    #         screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
    #         ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
    #         wb.save("test_results.xlsx")
    #     except AssertionError:
    #         result = "Fail"
    #         function_name = inspect.stack()[0][3]
    #         test_id = function_name.split("_")[-1]
    #         ws.append([f"TC-{test_id}", function_name, result])
    #
    # def test_orderdetails_button_ongoing_tc_06(self,appium_driver):
    #     try:
    #         main.TestAppiumAutomation.test_orderdetails_button(self,appium_driver)
    #         assert True
    #         result = "Pass"
    #     except TimeoutException:
    #         print("TimeoutException: Element not found within the specified timeout")
    #         result = "Fail"
    #         function_name = inspect.stack()[0][3]
    #         test_id = function_name.split("_")[-1]
    #         ws.append([f"TC-{test_id}", function_name, result])
=======
    def testorder_details_ongoing_tc_05(self,appium_driver):
        try:
            main.TestAppiumAutomation. test_dropdown_myorders_ongoing(self,appium_driver)
            time.sleep(7)
            screenshot_name = f"test_notification_tc_04.png"
            appium_driver.save_screenshot(screenshot_name)

            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
            wb.save("test_results.xlsx")
        except AssertionError:
            result = "Fail"
        function_name = inspect.stack()[0][3]
        test_id = function_name.split("_")[-1]
        ws.append([f"TC-{test_id}", function_name, result])

    def test_orderdetails_button_ongoing_tc_06(self,appium_driver):
        try:
            main.TestAppiumAutomation.test_orderdetails_button(self,appium_driver)
            assert True
            result = "Pass"
        except TimeoutException:
            print("TimeoutException: Element not found within the specified timeout")
            result = "Fail"
        function_name = inspect.stack()[0][3]
        test_id = function_name.split("_")[-1]
        ws.append([f"TC-{test_id}", function_name, result])
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
####################################################################################################################
    def test_payment_button_click_tc_05(self,appium_driver):
        try:
            main.TestAppiumAutomation. test_payments(self,appium_driver)
            time.sleep(7)
            screenshot_name = f"test_payment_tc_07.png"
            appium_driver.save_screenshot(screenshot_name)

            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append(
                [f"TC-{test_id}", function_name, result, screenshot_hyperlink, "None", "None", "payment button click",
                 "As Expected", "APK", "None"])
            wb.save("test_results.xlsx")
        except TimeoutException:
            result = "Fail"
            screenshot_name = f"test_payment_tc_07.png"
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
            wb.save("test_results.xlsx")
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "payment button click", "Not opening", "APK", "it's not clicking the app icon"])
###############################################################################################################################
    def test_performance_button_click_tc_06(self,appium_driver):
        try:
            main.TestAppiumAutomation. test_performance_button(self,appium_driver)
            time.sleep(7)
            screenshot_name = f"test_performance_tc_08.png"
            appium_driver.save_screenshot(screenshot_name)

            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append(
                [f"TC-{test_id}", function_name, result, screenshot_hyperlink, "None", "None", "performance button click",
                 "As Expected", "APK", "None"])
            wb.save("test_results.xlsx")
        except TimeoutException:
            result = "Fail"
            screenshot_name = f"test_performance_tc_08.png"
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
<<<<<<< HEAD

            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "performance button click", "Not opening", "APK", "it's not clicking the  icon"])
=======
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
            wb.save("test_results.xlsx")
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "performance button click", "Not opening", "APK", "it's not clicking the app icon"])
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
##################################################################################################################
    def test_account_button_click_tc_07(self,appium_driver):
        try:
            main.TestAppiumAutomation.test_account(self,appium_driver)
            time.sleep(7)
            screenshot_name = f"test_account_tc_09.png"
            appium_driver.save_screenshot(screenshot_name)

            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append(
                [f"TC-{test_id}", function_name, result, screenshot_hyperlink, "None", "None",
                 "Account button click",
                 "As Expected", "APK", "None"])
            wb.save("test_results.xlsx")
        except TimeoutException:
            result = "Fail"
            screenshot_name = f"test_account_tc_09.png"
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
<<<<<<< HEAD

=======
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
            wb.save("test_results.xlsx")
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "Account button click", "Not opening", "APK", "it's not clicking the app icon"])
<<<<<<< HEAD

=======
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
################################################################################################################
    def test_dutystatus_button_click_tc_08(self,appium_driver):
        try:
            main.TestAppiumAutomation.test_dutystatus_on(self,appium_driver)
            time.sleep(7)
            screenshot_name = f"test_statuson_tc_09.png"
            appium_driver.save_screenshot(screenshot_name)

            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append(
                [f"TC-{test_id}", function_name, result, screenshot_hyperlink, "None", "None",
                 "Duty status on and off",
                 "As Expected", "APK", "None"])
            wb.save("test_results.xlsx")
        except TimeoutException:
            result = "Fail"
            screenshot_name = f"test_statuson_tc_09.png"
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
            wb.save("test_results.xlsx")
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "Duty status on and off", "Not opening", "APK", "it's not clicking the app icon"])

########################################################################################################
    def test_ordercompleted_button_click_tc_09(self, appium_driver):
        try:
            main.TestAppiumAutomation.test_dropdown_myorders_completed(self, appium_driver)
            time.sleep(7)
            screenshot_name = f"test_completed_tc_09.png"
            appium_driver.save_screenshot(screenshot_name)
            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
<<<<<<< HEAD
            ws.append(
                [f"TC-{test_id}", function_name, result, screenshot_hyperlink, "None", "None",

                 "order complete button click", "APK", "None"])
=======
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
            wb.save("test_results.xlsx")
        except TimeoutException:
            result = "Fail"
            screenshot_name = f"test_completed_tc_09.png"
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
<<<<<<< HEAD

            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "order complete button click", "Not opening", "unable to select dropdown"])
=======
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
            wb.save("test_results.xlsx")
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            if __name__ == '__main__':
                ws.append([f"TC-{test_id}", function_name, result])
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
    ########################################################################################################
    def test_orderdetailscompleted_button_click_tc_10(self, appium_driver):
        try:
            main.TestAppiumAutomation.test_orderdetails_button_completed(self, appium_driver)
            time.sleep(7)
            screenshot_name = f"test_completed_tc_10.png"
            appium_driver.save_screenshot(screenshot_name)
            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
<<<<<<< HEAD
            ws.append(
                [f"TC-{test_id}", function_name, result, screenshot_hyperlink, "None", "None",
                 "order details completed"
                 "As Expected", "APK", "None"])
=======
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
            wb.save("test_results.xlsx")
        except TimeoutException:
            result = "Fail"
            screenshot_name = f"test_completed_tc_10.png"
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
<<<<<<< HEAD

            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "order details completed",
                       "Not opening", "APK", "order complete xpath"])
    #################################################################################################################
    def test_logout_button_click_TC_11(self,appium_driver):
        try:
            main.TestAppiumAutomation.test_logout_button(self,appium_driver)
            time.sleep(7)
            screenshot_name = f"test_completed_tc_11.png"
            appium_driver.save_screenshot(screenshot_name)
            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append(
                [f"TC-{test_id}", function_name, result, screenshot_hyperlink, "None", "None",
                 "logout button click",
                 "As Expected", "APK", "None"])
            wb.save("test_results.xlsx")

        except TimeoutException:
            result = "Fail"
            screenshot_name = f"test_completed_tc_10.png"
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
=======
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink])
            wb.save("test_results.xlsx")
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
<<<<<<< HEAD
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "logout button click", "Not opening", "APK", "it's not clicking the app icon"])


=======
            ws.append([f"TC-{test_id}", function_name, result])

    #################################################################################################################
    def test_logout_button_click(self,appium_driver):
        main.TestAppiumAutomation.test_logout_button(self,appium_driver)
>>>>>>> afe342d73554412cbc9855b2ceee7566124bedb9

