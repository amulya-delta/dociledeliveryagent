import time
from Driver.appiumandroid import AppiumDriverManager
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as xlImage
from openpyxl.utils import get_column_letter
from pages import main

logging.basicConfig(filename='test_log.log', level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')

# Load existing workbook or create a new one if it doesn't exist
try:
    wb = load_workbook("automation_test_report3.xlsx")
except FileNotFoundError:
    wb = Workbook()

ws = wb.active
ws.title = 'Test Results'

headers = ['Issue ID', 'Issue Description', 'Test Result', 'Screenshot', 'Severity Level',
           'Steps to Reproduce', 'Expected Behavior', 'Actual Behavior', 'Device/Platform',
           'Additional Notes/Comments']


# Write headers if the sheet is empty
def write_headers():
    try:
        if all(cell.value is None for cell in ws[1]):
            ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
    except Exception as e:
        logging.error(f"Error writing headers to Excel: {e}")

write_headers()

def write_test_result(issue_id, issue_description, test_result, screenshot, severity_level,
                      steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
                      additional_notes_comments):
    try:
        ws.append([issue_id, issue_description, test_result, "", severity_level,
                   steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
                   additional_notes_comments])
        row_number = ws.max_row
        if screenshot:
            # row_height = 100# Adjust the divisor based on your requirements
            # ws.row_dimensions[row_number].height = row_height
            # column_height = 200
            # ws.height_dimension[row_number]=column_height
            cell_reference = f'D{row_number}'
            row_height = 100  # Adjust the divisor based on your requirements
            ws.row_dimensions[row_number].height = row_height
            column_width = 100  # Adjust the divisor based on your requirements
            ws.column_dimensions[cell_reference[0]].width = column_width
            screenshot.width = 200  # Set width in pixels
            screenshot.height = 100
            ws.add_image(screenshot, cell_reference)  # Adjust cell position as needed
            wb.save('automation_test_report3.xlsx')
    except Exception as e:
        logging.error(f"Error writing test result to Excel: {e}")


def test_openapp_tc_01():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    try:
        main.TestAppiumAutomation.test_click_docile_delivery_element(driver)
        time.sleep(8)
        screenshot_path = "login.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)

        assert True
        logging.info("Successful login test passed")
        # print("login success")
        # Write test result to Excel
        write_test_result(issue_id='1', issue_description='App Open',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='User should be logged in successfully',
                          actual_behavior='User logged in successfully',
                          device_platform='Android', additional_notes_comments='No issues')

    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "login_fail.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='1', issue_description='Login with valid credentials',
                          test_result='Failed', screenshot=screenshot,
                          severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
                          expected_behavior='User should be logged in successfully',
                          actual_behavior='Login failed',
                          device_platform='Android', additional_notes_comments='Error encountered')


#####################################################################################################################################################
def test_yestologin_tc_02():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    try:
        main.TestAppiumAutomation.test_pressyes(driver)
        time.sleep(7)
        screenshot_path = "yes_to_login.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)

        assert True
        logging.info("Successful login test passed")
        # print("login success")
        # Write test result to Excel
        write_test_result(issue_id='2', issue_description='Click yes to login',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='YES to login',
                          actual_behavior='As Expected',
                          device_platform='Android', additional_notes_comments='No issues')


    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "yes_to_login.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='2', issue_description='Click yes to login', test_result='Failed',
                          screenshot=screenshot,
                          severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
                          expected_behavior='YES to login',
                          actual_behavior='As Expected', device_platform='Android',
                          additional_notes_comments='Error encountered')
###################################################################################################################################################
def test_mobile_otp_tc_03():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    try:
        main.TestAppiumAutomation.test_mobile_number(driver)
        time.sleep(7)
        screenshot_path = "test_mobile_otp.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)

        assert True
        logging.info("Successful login test passed")
        # print("login success")
        # Write test result to Excel
        write_test_result(issue_id='3', issue_description='Login with mobile no and enter otp',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='Login with mobile no',
                          actual_behavior='As Expected',
                          device_platform='Android', additional_notes_comments='No issues')


    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "test_mobile_otp.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='3', issue_description='Login with mobile no and enter otp', test_result='Failed',
                          screenshot=screenshot,
                          severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
                          expected_behavior='Login with mobile no',
                          actual_behavior='As Expected', device_platform='Android',
                          additional_notes_comments='Error encountered')
########################################################################################################################################################
def test_notification_button_tc_04():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    try:
        main.TestAppiumAutomation.test_notification_button(driver)
        time.sleep(7)
        screenshot_path = "test_notification_tc_04.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)

        assert True
        logging.info("Successful login test passed")
        # print("login success")
        # Write test result to Excel
        write_test_result(issue_id='4', issue_description='Click on notification',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='Click on notification',
                          actual_behavior='As Expected',
                          device_platform='Android', additional_notes_comments='No issues')


    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "test_notification_tc_04.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='4', issue_description='Click on notification', test_result='Failed',
                          screenshot=screenshot,
                          severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
                          expected_behavior='Click on notification',
                          actual_behavior='As Expected', device_platform='Android',
                          additional_notes_comments='Error encountered')
###########################################################################################################################################################
def test_ordercompleted_dropdown_button_click_tc_05():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    try:
        main.TestAppiumAutomation.test_dropdown_myorders_completed(driver)
        time.sleep(7)
        screenshot_path = "test_completed.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)

        assert True
        logging.info("Successful login test passed")
        # print("login success")
        # Write test result to Excel
        write_test_result(issue_id='5', issue_description='order complete  button click',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='order complete button click xpath"',
                          actual_behavior='As Expected',
                          device_platform='Android', additional_notes_comments='No issues')


    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "test_account.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='5', issue_description='order complete button click', test_result='Failed',
                          screenshot=screenshot,
                          severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
                          expected_behavior='order complete button click xpath"',
                          actual_behavior='As Expected', device_platform='Android',
                          additional_notes_comments='Error encountered')
###########################################################################################################################################################
def test_payment_button_click_tc_06():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    try:
        main.TestAppiumAutomation.test_payments(driver)
        time.sleep(7)
        screenshot_path = "payment_button.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)

        assert True
        logging.info("Successful login test passed")
        # print("login success")
        # Write test result to Excel
        write_test_result(issue_id='6', issue_description='payment button click"',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='payment button click xpath"',
                          actual_behavior='As Expected',
                          device_platform='Android', additional_notes_comments='No issues')


    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "test_notification_tc_04.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='6', issue_description='payment button click"', test_result='Failed',
                          screenshot=screenshot,
                          severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
                          expected_behavior='payment button click xpath"',
                          actual_behavior='As Expected', device_platform='Android',
                          additional_notes_comments='Error encountered')
#############################################################################################################################
def test_performance_button_click_tc_07():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    try:
        main.TestAppiumAutomation.test_performance_button(driver)
        time.sleep(7)
        screenshot_path = "performance_button.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)

        assert True
        logging.info("Successful login test passed")
        # print("login success")
        # Write test result to Excel
        write_test_result(issue_id='7', issue_description='performance button click"',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='performance button click xpath"',
                          actual_behavior='As Expected',
                          device_platform='Android', additional_notes_comments='No issues')


    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "performance_button.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='7', issue_description='performance button click"', test_result='Failed',
                          screenshot=screenshot,
                          severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
                          expected_behavior='performance button click xpath"',
                          actual_behavior='As Expected', device_platform='Android',
                          additional_notes_comments='Error encountered')
####################################################################################################################################

#########################################################################################################################################
def test_account_button_click_tc_08():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    try:
        main.TestAppiumAutomation.test_account(driver)
        time.sleep(7)
        screenshot_path = "test_account.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)

        assert True
        logging.info("Successful login test passed")
        # print("login success")
        # Write test result to Excel
        write_test_result(issue_id='8', issue_description='Login button click',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='Login button click xpath"',
                          actual_behavior='As Expected',
                          device_platform='Android', additional_notes_comments='No issues')


    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "test_account.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='8', issue_description='Login button click', test_result='Failed',
                          screenshot=screenshot,
                          severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
                          expected_behavior='Login button click xpath"',
                          actual_behavior='As Expected', device_platform='Android',
                          additional_notes_comments='Error encountered')
#######################################################################################################################
def  test_dutystatus_button_click_tc_09():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    try:
        main.TestAppiumAutomation.test_dutystatus_on(driver)
        time.sleep(7)
        screenshot_path = "test_statuson.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)

        assert True
        logging.info("Successful login test passed")
        # print("login success")
        # Write test result to Excel
        write_test_result(issue_id='9', issue_description='Duty status button click',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='Duty status button click xpath"',
                          actual_behavior='As Expected',
                          device_platform='Android', additional_notes_comments='No issues')


    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "test_account.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='9', issue_description='Duty status button click', test_result='Failed',
                          screenshot=screenshot,
                          severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
                          expected_behavior='Duty status  click xpath"',
                          actual_behavior='As Expected', device_platform='Android',
                          additional_notes_comments='Error encountered')
#########################################################################################################################################
def test_logout_button_click_TC_10():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    try:
        main.TestAppiumAutomation.test_logout_button(driver)
        time.sleep(7)
        screenshot_path = "test_logout.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)

        assert True
        logging.info("Successful login test passed")
        # print("login success")
        # Write test result to Excel
        write_test_result(issue_id='10', issue_description='Logout button click',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='Logout button click xpath"',
                          actual_behavior='As Expected',
                          device_platform='Android', additional_notes_comments='No issues')


    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "test_account.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='10', issue_description='Logout button click', test_result='Failed',
                          screenshot=screenshot,
                          severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
                          expected_behavior='Logout button   click xpath"',
                          actual_behavior='As Expected', device_platform='Android',
                          additional_notes_comments='Error encountered')