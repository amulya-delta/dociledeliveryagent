import time
import pytest
import inspect
from appium import webdriver
from appium.options.common.base import AppiumOptions
from pages import main
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


# Load the existing Excel workbook
wb = load_workbook(r"C:\Users\admin\PycharmProjects\Appium\test_results.xlsx")
ws = wb.active
# Define a fixture to save the workbook after the tests have run
@pytest.fixture(scope="session")
def excel_results(request):
    def fin():
        wb.save("test_results.xlsx")
        wb.close()

    request.addfinalizer(fin)
# Define a fixture to open the appium driver
# @pytest.fixture(scope="module")
# def appium_driver():
#     options = AppiumOptions()
#     options.load_capabilities({
#     "platformName": "Android",
#     "deviceName": "RZ8R21SMYNN",
#     })
#     driver = webdriver.Remote("http://127.0.0.1:4723/wd/hub", options=options)
#     yield driver
#     driver.quit()

@pytest.fixture(scope="module")
def appium_driver():
    options = AppiumOptions()
    options.load_capabilities({
        "platformName": "Android",
        "deviceName": "RZ8R21SMYNN",
        # Add other desired capabilities as needed
    })
    # Add other desired capabilities as needed

    driver = webdriver.Remote("http://127.0.0.1:4723/wd/hub", options=options)
    yield driver
    # driver.quit()


@pytest.mark.usefixtures("appium_driver","excel_results")
class Testorderpage:
    @staticmethod
    def test_openapp_tc_01(appium_driver, excel_results):
        try:

            main.TestAppiumAutomation.test_click_docile_delivery_element(appium_driver)
            time.sleep(7)
            screenshot_name = f"test_openapp_tc_01.png"
            # allure.attach(appium_driver.get_screenshot_as_png(), name="Screenshot",
            #               attachment_type=allure.attachment_type.PNG)

            appium_driver.save_screenshot(screenshot_name)
            assert True  # Example assertion
            result = "Pass"
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.cell(row=2, column=1)
            ws.append(
                [f"TC-{test_id}", function_name, result, "", "None", "None", "open the docile app",
                 "As Expected", "APK", "None"])


            img = Image(screenshot_name)
            img.width = 3 * 96 * 0.393701  # Convert cm to pixels (assuming 96 DPI)
            img.height = 6 * 96 * 0.393701
            ws.add_image(img, f'D{ws.max_row}')
            wb.save("test_results.xlsx")

        except TimeoutException:
            result = "Fail"
            screenshot_name = f"test_openapp_tc_01.png"
            # allure.attach(appium_driver.get_screenshot_as_png(), name="Screenshot",
            #               attachment_type=allure.attachment_type.PNG)
            appium_driver.save_screenshot(screenshot_name)
            function_name = inspect.stack()[0][3]
            test_id = function_name.split("_")[-1]
            screenshot_hyperlink = f'=HYPERLINK("{screenshot_name}", "Screenshot")'
            ws.append([f"TC-{test_id}", function_name, result, screenshot_hyperlink, "High", "Try using ID or Xpath",
                       "open docile App", "Not opening", "APK", "it's not clicking the app icon"])
            wb.save("test_results.xlsx")